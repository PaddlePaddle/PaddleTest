"""
get benchmark info from log
"""

import os
import sys
import json
import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

import base_mkldnn_fp32
import base_mkldnn_int8
import base_trt_fp16
import base_trt_int8
import base_nv_trt_fp16
import base_nv_trt_int8
import mail_report
import write_db


FONT = {
    "g": "00ff00",
    "b": "ff0000",
    "s": "000000",
    "o": "cccccc",
}


def get_runtime_info(log_file):
    """
    获取本次执行结果
    """
    benchmark_res = {}
    if not os.path.exists(log_file):
        return benchmark_res
    with open(log_file) as fin:
        benchmark_lines = ""
        lines = fin.readlines()
        for line in lines:
            if "[Benchmark][final result]" in line:
                tmp = line.strip("[Benchmark][final result]").strip()
                res_json = eval(tmp)
                model_name = res_json["model_name"]
                benchmark_res[model_name] = res_json
    return benchmark_res


def get_base_info(mode):
    """
    从base文件中读取base数据
    mode: trt_int8 trt_fp16 mkldnn_int8 mkldnn_fp32
    """
    if mode == "trt_int8":
        base_res = base_trt_int8.trt_int8
    elif mode == "trt_fp16":
        base_res = base_trt_fp16.trt_fp16
    elif mode == "nv_trt_int8":
        base_res = base_nv_trt_int8.nv_trt_int8
    elif mode == "nv_trt_fp16":
        base_res = base_nv_trt_fp16.nv_trt_fp16
    elif mode == "mkldnn_int8":
        base_res = base_mkldnn_int8.mkldnn_int8
    elif mode == "mkldnn_fp32":
        base_res = base_mkldnn_fp32.mkldnn_fp32
    else:
        base_res = None
    return base_res


def compare_diff(base_res, benchmark_res, metric_list):
    """
    计算本次结果与base的diff、gsb
    """
    benchmark_keys = benchmark_res.keys()
    compare_res = {}
    for model, info in base_res.items():
        compare_res[model] = {}
        for item in metric_list:
            compare_res[model][item] = {
                "th": info[item]["th"],
                "base": info[item]["value"],
                "benchmark": -1,
                "diff": -1,
                "gsb": "o",
                "unit": info[item]["unit"],
            }

        if model not in benchmark_keys:
            continue

        for item in metric_list:
            compare_res[model][item]["benchmark"] = benchmark_res[model][item]["value"]
            if compare_res[model][item]["base"] <= 0:
                continue
            gap = compare_res[model][item]["benchmark"] - compare_res[model][item]["base"]
            diff = gap / compare_res[model][item]["base"]
            compare_res[model][item]["diff"] = diff
            if diff >= -compare_res[model][item]["th"] and diff <= compare_res[model][item]["th"]:
                compare_res[model][item]["gsb"] = "s"
            elif diff < -compare_res[model][item]["th"]:
                compare_res[model][item]["gsb"] = "b"
            elif diff > compare_res[model][item]["th"]:
                compare_res[model][item]["gsb"] = "g"

    return compare_res


def gsb(compare_res, metric_list):
    """
    统计compare_res的gsb
    """
    gsb = {}
    for item in metric_list:
        gsb.setdefault(
            item,
            {
                "gsb": "",
                "g": 0,
                "s": 0,
                "b": 0,
                "total": 0,
                "b_ratio": 0,
                "g_ratio": 0,
            },
        )
    for model, info in compare_res.items():
        for item in metric_list:
            if info[item]["gsb"] == "g":
                gsb[item]["g"] += 1
            elif info[item]["gsb"] == "s":
                gsb[item]["s"] += 1
            elif info[item]["gsb"] == "b":
                gsb[item]["b"] += 1
            gsb[item]["gsb"] = "{}:{}:{}".format(gsb[item]["g"], gsb[item]["s"], gsb[item]["b"])
            gsb[item]["total"] = gsb[item]["g"] + gsb[item]["s"] + gsb[item]["b"]
            if gsb[item]["total"] > 0:
                gsb[item]["b_ratio"] = gsb[item]["b"] / gsb[item]["total"]
                gsb[item]["g_ratio"] = gsb[item]["g"] / gsb[item]["total"]

    return gsb


def res_summary(diff_res, mode_list, metric_list):
    """
    汇总不同模式下的数据
    """
    tongji = {}
    res = {}

    gsb_total = {}
    for item in metric_list:
        gsb_total.setdefault(
            item,
            {
                "gsb": "",
                "g": 0,
                "s": 0,
                "b": 0,
                "total": 0,
                "b_ratio": 0,
                "g_ratio": 0,
            },
        )

    # 统计数据
    for mode in mode_list:
        _gsb = gsb(diff_res[mode], metric_list)
        tongji.setdefault(mode, _gsb)

    for item in metric_list:
        for mode in mode_list:
            gsb_total[item]["g"] += tongji[mode][item]["g"]
            gsb_total[item]["s"] += tongji[mode][item]["s"]
            gsb_total[item]["b"] += tongji[mode][item]["b"]
            gsb_total[item]["total"] += tongji[mode][item]["total"]
        gsb_total[item]["gsb"] = "{}:{}:{}".format(gsb_total[item]["g"], gsb_total[item]["s"], gsb_total[item]["b"])
        if gsb_total[item]["total"] > 0:
            gsb_total[item]["b_ratio"] = gsb_total[item]["b"] / gsb_total[item]["total"]
            gsb_total[item]["g_ratio"] = gsb_total[item]["g"] / gsb_total[item]["total"]

    tongji.setdefault("total", gsb_total)

    # 详细数据
    m = list(diff_res.keys())
    if len(m) < 1:
        return res, tongji
    models = diff_res[m[0]].keys()
    for model in models:
        res[model] = {}
        for mode in mode_list:
            res[model][mode] = {}
            for item in metric_list:
                res[model][mode][item] = {
                    "th": diff_res[mode][model][item]["th"],
                    "base": diff_res[mode][model][item]["base"],
                    "benchmark": diff_res[mode][model][item]["benchmark"],
                    "diff": diff_res[mode][model][item]["diff"],
                    "gsb": diff_res[mode][model][item]["gsb"],
                    "unit": diff_res[mode][model][item]["unit"],
                }

    return res, tongji


def res2xls(env, res, tongji, mode_list, metric_list, jingping_list, save_file):
    """
    将结果保存为excel文件
    """

    wb = openpyxl.Workbook()

    # table1: 详细数据
    sheet_detail = wb.create_sheet(index=0, title="detail")
    sheet_detail = wb["detail"]

    # 表头
    sheet_detail.cell(1, 2).value = env

    D = 1 + 2 * len(jingping_list)
    n = len(metric_list) * D
    row_s = 2
    column_s = 2
    for m in mode_list:
        sheet_detail.cell(row_s, column_s).value = m
        column_s += n

    row_s = 3
    column_s = 2
    for m in mode_list:
        for k in metric_list:
            sheet_detail.cell(row_s, column_s).value = k
            column_s += D

    row_s = 4
    column_s = 2
    n = len(mode_list) * len(metric_list)
    for i in range(n):
        w = 0
        sheet_detail.cell(4, column_s + w).value = "实际值"
        for jingping in jingping_list:
            w += 1
            sheet_detail.cell(4, column_s + w).value = "{}值".format(jingping)
            w += 1
            sheet_detail.cell(4, column_s + w).value = "diff-{}".format(jingping)
        column_s += D

    # 数据
    row_s = 5
    models = list(res["base"].keys())
    for model in models:
        column_s = 1
        w = 0
        sheet_detail.cell(row_s, column_s).value = model
        for m in mode_list:
            for k in metric_list:
                w += 1
                sheet_detail.cell(row_s, column_s + w).value = res["base"][model][m][k]["benchmark"]
                for jingping in jingping_list:
                    if (model in res[jingping].keys()) and (m in res[jingping][model].keys()):
                        w += 1
                        sheet_detail.cell(row_s, column_s + w).value = res[jingping][model][m][k]["base"]
                        w += 1
                        sheet_detail.cell(row_s, column_s + w).value = res[jingping][model][m][k]["diff"]
                        _color = res[jingping][model][m][k]["gsb"]
                        _font = Font(color=FONT[_color])
                        sheet_detail.cell(row_s, column_s + w).font = _font
                    else:
                        w += 1
                        w += 1
        row_s += 1

    # 合并表头单元格
    """
    sheet_detail.merge_cells("B1:Z1")
    sheet_detail.merge_cells("B2:I2")
    sheet_detail.merge_cells("B3:E3")
    sheet_detail.merge_cells("F3:I3")
    sheet_detail.merge_cells("J2:Q2")
    sheet_detail.merge_cells("J3:M3")
    sheet_detail.merge_cells("N3:Q3")
    sheet_detail.merge_cells("R2:Y2")
    sheet_detail.merge_cells("R3:U3")
    sheet_detail.merge_cells("V3:Y3")
    sheet_detail.merge_cells("Z2:AG2")
    sheet_detail.merge_cells("Z3:AC3")
    sheet_detail.merge_cells("AD3:AG3")
    """

    # table2: gsb统计数据
    for jingping in jingping_list:
        add_table_gsb(wb, jingping, tongji[jingping], metric_list, mode_list)

    wb.save("{}".format(save_file))


def add_table_gsb(wb, jingping, tongji, metric_list, mode_list):
    """
    gsb
    """
    title = "gsb-{}".format(jingping)
    sheet_gsb = wb.create_sheet(index=1, title=title)
    sheet_gsb = wb[title]

    # 表头
    column_s = 2
    row_s = 1
    for k in metric_list:
        sheet_gsb.cell(row_s, column_s).value = k
        column_s += 3

    column_s = 2
    row_s = 2
    for k in metric_list:
        sheet_gsb.cell(row_s, column_s).value = "GSB"
        sheet_gsb.cell(row_s, column_s + 1).value = "下降数（占比）"
        sheet_gsb.cell(row_s, column_s + 2).value = "上升数（占比）"
        column_s += 3

    # 数据
    column_s = 1
    row_s = 3
    for m in mode_list:
        info = tongji[m]
        sheet_gsb.cell(row_s, column_s).value = m
        for k in metric_list:
            sheet_gsb.cell(row_s, column_s + 1).value = info[k]["gsb"]
            sheet_gsb.cell(row_s, column_s + 2).value = "{}({})".format(info[k]["b"], info[k]["b_ratio"])
            sheet_gsb.cell(row_s, column_s + 3).value = "{}({})".format(info[k]["g"], info[k]["g_ratio"])
            column_s += 3
        column_s = 1
        row_s += 1

    # 并表头单元格
    sheet_gsb.merge_cells("B1:D1")
    sheet_gsb.merge_cells("E1:G1")


def res2db(env, benchmark_res, mode_list, metric_list):
    """
    转化为db需要的数据格式，部分字段取值待定
    """
    res = []
    for mode in mode_list:
        for model, info in benchmark_res[mode].items():
            item = {
                "task_dt": env["task_dt"],
                "model_name": model,
                "batch_size": info["batch_size"],
                "fp_mode": "int8",
                "use_trt": True,
                "use_mkldnn": False,
                "jingdu": info["jingdu"]["value"],
                "jingdu_unit": info["jingdu"]["unit"],
                "ips": info["xingneng"]["value"],
                "ips_unit": info["xingneng"]["unit"],
                "cpu_mem": info["cpu_mem"]["value"],
                "gpu_mem": info["gpu_mem"]["value"],
                "frame": env["frame"],
                "frame_branch": env["frame_branch"],
                "frame_commit": env["frame_commit"],
                "frame_version": env["frame_version"],
                "docker_image": env["docker_image"],
                "python_version": env["python_version"],
                "cuda_version": env["cuda_version"],
                "cudnn_version": env["cudnn_version"],
                "trt_version": env["trt_version"],
                "device": env["device"],
                "thread_num": 1,
            }
            res.append(item)
    return res


def run():
    """
    统计结果并保存到excel文件
    """
    task_dt = datetime.date.today()

    frame = sys.argv[1]
    frame_branch = sys.argv[2]
    frame_commit = sys.argv[3]
    frame_version = sys.argv[4]
    docker_image = sys.argv[5]
    python_version = sys.argv[6]
    cuda_version = sys.argv[7]
    cudnn_version = sys.argv[8]
    trt_version = sys.argv[9]
    device = sys.argv[10]
    modes = sys.argv[11]
    metrics = sys.argv[12]
    save_file = sys.argv[13]

    mode_list = modes.split(",")
    metric_list = metrics.split(",")

    env = {
        "task_dt": task_dt,
        "frame": frame,
        "frame_branch": frame_branch,
        "frame_commit": frame_commit,
        "frame_version": frame_version,
        "docker_image": docker_image,
        "python_version": python_version,
        "cuda_version": cuda_version,
        "cudnn_version": cudnn_version,
        "trt_version": trt_version,
        "device": device,
        "threshold": "时延/内存/显存 0.05，精度 0.01",
    }

    benchmark_res = {}
    diff_res = {}
    diff_res_nv = {}
    for mode in mode_list:
        log_file = "eval_{}_acc.log".format(mode)
        _current = get_runtime_info(log_file)
        benchmark_res.setdefault(mode, _current)
        _base = get_base_info(mode)
        _diff = compare_diff(_base, _current, metric_list)
        diff_res.setdefault(mode, _diff)
        if mode in ["trt_int8", "trt_fp16"]:
            _base_nv = get_base_info("nv_" + mode)
            _diff_nv = compare_diff(_base_nv, _current, metric_list)
            diff_res_nv.setdefault(mode, _diff_nv)

    res_base, tongji_base = res_summary(diff_res, mode_list, metric_list)
    res_nv, tongji_nv = res_summary(diff_res_nv, list(diff_res_nv.keys()), metric_list)
    res = {
        "base": res_base,
        "NV-TRT": res_nv,
    }
    tongji = {
        "base": tongji_base,
        "NV-TRT": tongji_nv,
    }
    jingping_list = ["base"]
    if "trt_int8" in mode_list:
        jingping_list.append("NV-TRT")

    env_str = "环境: "
    env_str += "docker: "
    env_str += docker_image
    env_str += "  "
    env_str += "frame: "
    env_str += "  "
    env_str += "frame_branch: "
    env_str += frame_branch
    env_str += "  "
    env_str += "frame_commit: "
    env_str += frame_commit
    env_str += "  "
    env_str += "device: "
    env_str += gpu
    env_str += "  "
    env_str += cpu
    env_str += "  "
    env_str += "阈值: 时延/内存/显存 0.05，精度 0.01"
    env_str += "  "

    # save result to xlsx
    res2xls(env_str, res, tongji, mode_list, metric_list, jingping_list, save_file)

    # save result to db
    db_res = res2db(env, benchmark_res, mode_list, metric_list)
    write_db.write(db_res)

    # send mail
    mail_report.report_day(task_dt, env, tongji, res, mode_list, metric_list, jingping_list)


if __name__ == "__main__":
    run()
