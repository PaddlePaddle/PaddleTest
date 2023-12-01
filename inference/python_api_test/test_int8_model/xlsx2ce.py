"""
# Copyright (c) 2020 PaddlePaddle Authors. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""


import os
import json
import argparse
import smtplib
import time
import logging as logger

from email.mime.text import MIMEText
from email.header import Header
from openpyxl import load_workbook
from openpyxl.styles import *
import yaml
import pandas as pd
import numpy as np


def parse_args():
    """
    parse input args
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--now_excel_name", type=str, default="tipc_benchmark_paddle.xlsx", help="output excel file name"
    )
    parser.add_argument("--base_yaml", type=str, default="xpu_int8_base.yaml", help="base yaml file name")
    parser.add_argument("--branch", type=str, default="develop", help="output excel file name")
    parser.add_argument("--no_send", type=str, default="False", help="debug mode receiver")

    return parser.parse_args()


def read_excel(padle_res_excel):
    """
    set excel style
    """
    res_mess = []
    workbook = load_workbook(padle_res_excel)
    sheet = workbook["Sheet1"]
    maxrows = sheet.max_row
    maxcolumns = sheet.max_column
    for row in range(2, maxrows + 1):
        temp_mess = []
        for col in range(1, maxcolumns + 1):
            temp_mess.append(sheet.cell(row, col).value)
            # res_json = kpi_params(temp_mess)
        res_mess.append(temp_mess)
    return sorted(res_mess)

    # send(url=url,res_json=res_json)


def calculate_gsb(df, threshold=5):
    """
    calculate gsb func
    """
    gsb = (
        f"{len(df[df['avg_cost_diff(%)'] > threshold])}:"
        f"{len(df[(df['avg_cost_diff(%)'] >= -threshold) & (df['avg_cost_diff(%)'] <= threshold)])}:"
        f"{len(df[df['avg_cost_diff(%)'] < -threshold])}"
    )
    return gsb


def gsb_function(data_df):
    """
    gsb msg func
    """
    # calculate
    good_count = len(data_df[data_df["avg_cost_diff(%)"] > 10])
    bad_count = len(data_df[data_df["avg_cost_diff(%)"] < -10])
    total_count = len(data_df)
    diff_10_df = data_df[data_df["avg_cost_diff(%)"] < -10]
    diff_30_df = data_df[data_df["avg_cost_diff(%)"] < -30]

    # important data
    error_df = data_df[(abs(data_df["avg_cost"]) == np.inf) | (pd.isnull(data_df["avg_cost"]))]
    normal_df = data_df[(abs(data_df["avg_cost_diff(%)"]) != np.inf) & (pd.notnull(data_df["avg_cost_diff(%)"]))]
    normal_count = len(normal_df)
    GSB = (good_count - bad_count) / normal_count * 100
    big_diff_df = data_df[abs(data_df["avg_cost_diff(%)"]) > 5]
    important_df = pd.concat([big_diff_df, error_df], ignore_index=True)

    # GSB
    total_gsb = calculate_gsb(normal_df, 5)

    msg = "<h3>【汇总信息】：<br>"
    msg += f"模型数：{data_df['model_name'].drop_duplicates().count()}<br>"
    msg += f"case总数：{total_count}<br>"
    msg += f"报错case数：{len(error_df)}<br>"
    msg += f"综合GSB：{total_gsb}<br>"
    msg += f"性能下降10%以上的case数量：{len(diff_10_df)} case占比：{round(len(diff_10_df) / normal_count * 100, 3)}%<br>"
    msg += f"性能下降30%以上的case数量：{len(diff_30_df)} case占比：{round(len(diff_30_df) / normal_count * 100, 3)}%</h3>"

    msg += "<h3>一、异常数据<br>"
    msg += f" \
                    <table style='display:DISPLAY;' border='1' align=center> \
                        <caption bgcolor='#989898'>异常数据</caption> \
                        <tr bgcolor='#989898'>"

    for column in list(important_df[:0]):
        msg += f"<td>{column}</td>"

    msg += "</tr>"

    for i, row in important_df.iterrows():
        msg += "<tr>"
        for index in row.index:
            bgcolor = "Cornsilk"
            if index == "model_name":
                bgcolor = "white"
            if "diff(%)" in index:
                # diff列
                if round(row[index], 3) < -5 or pd.isnull(row[index]):
                    bgcolor = "#FF6666"
                elif round(row[index], 3) > 5:
                    bgcolor = "#66FF66"
            if index in ["精度", "精度_base"]:
                td_data = round(row[index], 5)
            else:
                td_data = round(row[index], 3) if isinstance(row[index], float) else row[index]
            msg += f"<td bgcolor={bgcolor}>{td_data}</td>"
        msg += "</tr>"
    msg += "</table>"

    # for column in diff_columns:
    msg += "<h3>二、全量数据<br>"
    msg += f" \
                <table style='display:DISPLAY;' border='1' align=center> \
                    <caption bgcolor='#989898'>全量数据</caption> \
                    <tr bgcolor='#989898'>"

    for column in list(data_df[:0]):
        msg += f"<td>{column}</td>"

    msg += "</tr>"

    for i, row in data_df.iterrows():
        # print(row["model_name"])
        msg += "<tr>"
        for index in row.index:
            bgcolor = "Cornsilk"
            if index == "model_name":
                bgcolor = "white"
            if "diff(%)" in index:
                # diff列
                if round(row[index], 3) < -5 or pd.isnull(row[index]):
                    bgcolor = "#FF6666"
                elif round(row[index], 3) > 5:
                    bgcolor = "#66FF66"
            if index in ["精度", "精度_base"]:
                td_data = round(row[index], 5)
            else:
                td_data = round(row[index], 3) if isinstance(row[index], float) else row[index]
            msg += f"<td bgcolor={bgcolor}>{td_data}</td>"
        msg += "</tr>"
    msg += "</table>"

    return msg


def mail_msg(res_json_list, device_name, env):
    """
    main mail msg func
    """
    # 定义需要展示的列和需要比较diff的列
    all_columns = [
        "日期",
        "环境",
        "version",
        "device_name",
        "model_name",
        "repo",
        "precision",
        "l3_cache",
        "unit",
        "精度",
        "avg_cost",
        "HBM_used",
    ]
    select_columns = ["model_name", "repo", "precision", "l3_cache", "unit", "精度", "avg_cost", "HBM_used"]
    diff_columns = ["精度", "avg_cost", "HBM_used"]
    head_columns = ["model_name", "repo", "precision", "l3_cache", "unit"]
    now_df = pd.read_excel(args.now_excel_name)[select_columns]
    # develop_df = pd.read_excel(args.develop_excel_name)[select_columns]
    # base值
    base_df = pd.DataFrame(columns=all_columns)
    data_dict = yaml.load(open(args.base_yaml, "r", encoding="utf-8"), Loader=yaml.FullLoader)
    for k, v in data_dict.items():
        v["model_name"] = k
        base_df = pd.concat([base_df, pd.Series(v).to_frame().T], ignore_index=True)
    origin_df = pd.read_excel(args.now_excel_name).drop("Unnamed: 0", axis=1)
    # 相同config保留最后一次数据，重跑机制决定最后一次数据(最优或首次diff<5%)
    now_df = now_df.drop_duplicates(head_columns, keep="last")
    # develop_df = develop_df.drop_duplicates(head_columns, keep="last")
    origin_df.to_excel("tipc_benchmark_paddle_xpu_all.xlsx")
    origin_df = origin_df.drop_duplicates(head_columns, keep="last")
    origin_df.to_excel("tipc_benchmark_paddle_xpu.xlsx")

    merge_df_dev = pd.merge(base_df, now_df, how="inner", on=head_columns, suffixes=("_base", ""))

    merge_df_dev["avg_cost"] = merge_df_dev["avg_cost"].replace(0, np.nan)
    for col in diff_columns:
        merge_df_dev[f"{col}_diff(%)"] = merge_df_dev[[f"{col}_base", f"{col}"]].apply(
            lambda x: (x[f"{col}_base"] - x[f"{col}"]) / x[f"{col}_base"] * 100, axis=1
        )

    merge_df_dev.sort_values(by=head_columns, inplace=True)
    # re sort columns
    merge_columns = head_columns
    for col in diff_columns:
        merge_columns.append(f"{col}_base")
        merge_columns.append(col)
        merge_columns.append(f"{col}_diff(%)")
    # # TODO：精度比较阈值
    # merge_columns.append("diff_1e-2")
    merge_df_dev = merge_df_dev[merge_columns]
    merge_df_dev.sort_values(by=["repo", "model_name", "precision", "l3_cache"], inplace=True)

    merge_df_dev.to_excel("benchmark_dev.xlsx")
    # print(merge_df_dev)

    global cur_date
    cur_date = time.strftime("%Y-%m-%d", time.localtime(time.time()))

    # 生成diff case, diff率大于5就rerun
    big_diff_df = merge_df_dev[abs(merge_df_dev["avg_cost_diff(%)"]) > 5]
    big_diff_df.to_excel("benchmark_diff.xlsx")
    diff_case = []
    for i, row in big_diff_df.iterrows():
        print(i, row)
        name = row["model_name"]
        line = f"{name} {row['l3_cache']} {row['precision']} {row['avg_cost_base']} {row['avg_cost_diff(%)']}\n"
        diff_case.append(line)
    with open("diff_case.txt", "w", encoding="utf-8") as f:
        f.writelines(diff_case)

    monitor_msg = gsb_function(merge_df_dev)

    # 标题
    msg = "<meta http-equiv='Content-Type' content='text/html; charset=utf-8'> "
    msg += "<h3>【相关配置】：<br>"
    msg += "镜像：" + env + "<br>"
    msg += "测试硬件：" + device_name + "<br>"
    msg += "config配置：[XPU int8] * [关闭L3 Cache]<br>"
    msg += "<h3>【commit信息】：<br>"
    # msg += "上次commit：" + latest_commit + "<br>"
    msg += "本次commit：<br>"
    msg += this_commit + "</h3>"
    msg += monitor_msg

    return msg


def send_report(mail_msg):
    """
    send report func
    """
    sender = os.environ.get("sender")
    receiver_list = eval(os.environ.get("receiver_list"))
    print(sender)
    print(receiver_list)

    # message = MIMEMultipart()
    mail_msg = mail_msg
    msg = MIMEText(mail_msg, "html", "utf-8")
    cur_date = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    subject = f"【benchmark XPU】【Paddle Inference】{branch}分支报告"
    msg["From"] = sender
    msg["To"] = ",".join(receiver_list)
    msg["Subject"] = Header(subject, "utf-8")
    try:
        smtpObj = smtplib.SMTP("mail2-in.baidu.com", 25)
        smtpObj.sendmail(sender, receiver_list, msg.as_string())
        logger.info("Send mail success, " + cur_date)
        print("Send mail Success, " + cur_date)
        smtpObj.quit()
    except smtplib.SMTPException as e:
        logger.error("Send mail has exception: %s" % e)
        logger.error("Error: Send mail fail")
        print("Send mail has exception: %s" % e)
        print("Error: Send mail fail")


if __name__ == "__main__":
    args = parse_args()
    branch = args.branch

    now_res_mess_list = read_excel(args.now_excel_name)
    # develop_res_mess_list = read_excel(args.develop_excel_name)
    res_json_list = []
    env = now_res_mess_list[0][2]
    # latest_commit = str(develop_res_mess_list[0][3]).split("/")[1]
    this_commit = str(now_res_mess_list[0][3]).split("/")[1]
    device = str(now_res_mess_list[0][4])

    msg = mail_msg(res_json_list, device, env)  # 邮件数据格式转换
    if args.no_send == "False":
        send_report(msg)  # 发邮件
