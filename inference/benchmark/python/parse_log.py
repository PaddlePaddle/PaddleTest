# -*- coding: utf-8 -*-
# encoding=utf-8 vi:ts=4:sw=4:expandtab:ft=python
"""
parse benchmark log to excel
"""

import os
import re
import argparse

from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
import pandas as pd


def parse_args():
    """
    parse input args
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--log_path", type=str, default="./logs", help="benchmark log path")
    parser.add_argument("--output_name", type=str, default="benchmark_excel.xlsx", help="output excel file name")
    parser.add_argument("--process_trt", dest="process_trt", action="store_true")
    return parser.parse_args()


def find_all_logs(path_walk: str):
    """
    find all .log files from target dir
    """
    for root, ds, files in os.walk(path_walk):
        for file_name in files:
            if re.match(r".*.log", file_name):
                full_path = os.path.join(root, file_name)
                yield file_name, full_path


def process_log(file_name: str) -> dict:
    """
    process log to dict
    """
    output_dict = {}
    with open(file_name, "r") as f:
        try:
            for i, data in enumerate(f.readlines()):
                if i == 0:
                    continue
                line_lists = data.split(" ")
                if "name:" in line_lists and "type:" in line_lists:
                    pos_buf = line_lists.index("name:")
                    output_dict["model_name"] = line_lists[pos_buf + 1].split(",")[0]
                    output_dict["frame_work"] = line_lists[-1].strip()
                if "Num" in line_lists and "size:" in line_lists:
                    pos_buf = line_lists.index("size:")
                    output_dict["batch_size"] = line_lists[pos_buf + 1].split(",")[0]
                if "device:" in line_lists:
                    pos_buf = line_lists.index("device:")
                    output_dict["device"] = line_lists[pos_buf + 1].strip()
                if "QPS:" in line_lists and "latency(ms):" in line_lists:
                    pos_buf = line_lists.index("QPS:")
                    output_dict["Average_latency(ms)"] = line_lists[pos_buf - 1].split(",")[0]
                    output_dict["QPS"] = line_lists[-1].strip()
                if "cpu_math_library_num_threads:" in line_lists:
                    output_dict["cpu_math_library_num_threads"] = line_lists[-1].strip()
                if "trt_precision:" in line_lists:
                    output_dict["trt_precision"] = line_lists[-1].strip()
        except Exception:
            output_dict["model_name"] = file_name

    return output_dict


def set_style(diff_excel):
    """
    set excel style
    """
    workbook = load_workbook(diff_excel)
    sheet1 = workbook.active
    cells = sheet1["A:J"]
    # center
    aligncenter = Alignment(horizontal="center", vertical="center")
    for i in cells:
        for j in i:
            j.alignment = aligncenter

    workbook.save(diff_excel)
    # edit width
    for index, i in enumerate(cells):
        column_width = 0
        for j in i:
            if j.value:
                if len(str(j.value)) > column_width:
                    column_width = len(str(j.value)) + 2
        sheet1.column_dimensions[get_column_letter(index + 1)].width = column_width

    workbook.save(diff_excel)


def main():
    """
    main
    """
    args = parse_args()
    # create empty DataFrame
    origin_df = pd.DataFrame(
        columns=["frame_work", "model_name", "batch_size", "device", "trt_precision", "Average_latency(ms)", "QPS"]
    )

    for file_name, full_path in find_all_logs(args.log_path):
        dict_log = process_log(full_path)
        origin_df = origin_df.append(dict_log, ignore_index=True)

    raw_df = origin_df.sort_values(by=["frame_work", "model_name", "batch_size", "device", "trt_precision"])
    raw_df.to_excel(args.output_name)
    set_style(args.output_name)


if __name__ == "__main__":
    main()
